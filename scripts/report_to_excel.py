import pandas as pd
from openpyxl import load_workbook

def append_to_excel(df, file_path, sheet_name):
    try:
        # Try to open the existing workbook
        book = load_workbook(file_path)
    except FileNotFoundError:
        # If the file doesn't exist, create a new one
        df.to_excel(file_path, sheet_name=sheet_name, index=False)
        return

    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl') as writer:
        if sheet_name in book.sheetnames:
            startrow = book[sheet_name].max_row  # Append below existing data
        else:
            startrow = 0  # Create new sheet

        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=(startrow == 0))

    print(f"Appended data to '{sheet_name}' in {file_path}")

# Example usage:
df_new_vendors = pd.DataFrame({"Action": ["New Vendor Added"], "Vendor Name": ["Vendor A"]})
df_updated_vendors = pd.DataFrame({"Action": ["Vendor Name Updated"], "Vendor Name": ["Vendor B"]})

append_to_excel(df_new_vendors, "vendors_report.xlsx", "Vendors")
append_to_excel(df_updated_vendors, "vendors_report.xlsx", "Updated Vendors")
